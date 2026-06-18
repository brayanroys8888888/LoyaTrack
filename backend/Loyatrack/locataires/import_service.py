"""Import en masse de locataires depuis un fichier CSV ou Excel — module 2.4."""
import csv
import io
from datetime import datetime
from decimal import Decimal, InvalidOperation

from .models import Locataire

COLONNES = [
    'nom', 'prenom', 'telephone', 'logement', 'montant_loyer',
    'jour_echeance', 'date_entree', 'solde_initial', 'date_debut_facturation',
]


def modele_csv():
    """Renvoie le contenu d'un template CSV téléchargeable (en-têtes)."""
    out = io.StringIO()
    csv.writer(out).writerow(COLONNES)
    return out.getvalue()


def _parse_date(valeur):
    if not valeur:
        return None
    valeur = str(valeur).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(valeur, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"date invalide: {valeur}")


def _lignes_depuis_fichier(contenu_bytes, nom_fichier):
    """Renvoie une liste de dicts {colonne: valeur} depuis un CSV ou XLSX."""
    if nom_fichier.lower().endswith(('.xlsx', '.xls')):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(contenu_bytes), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        entetes = [str(c).strip().lower() if c is not None else '' for c in rows[0]]
        return [dict(zip(entetes, r)) for r in rows[1:]]
    # CSV
    texte = contenu_bytes.decode('utf-8-sig')
    return list(csv.DictReader(io.StringIO(texte)))


def importer_locataires(bailleur, contenu_bytes, nom_fichier, dry_run=False):
    """Importe les locataires. Renvoie {crees, erreurs:[{ligne, message}], apercu:[...]}"""
    lignes = _lignes_depuis_fichier(contenu_bytes, nom_fichier)
    crees, erreurs, apercu = 0, [], []

    for i, row in enumerate(lignes, start=2):  # ligne 1 = en-têtes
        row = {(k or '').strip().lower(): v for k, v in row.items()}
        try:
            nom = (str(row.get('nom') or '')).strip()
            prenom = (str(row.get('prenom') or '')).strip()
            if not nom or not prenom:
                raise ValueError("nom et prenom obligatoires")
            montant = Decimal(str(row.get('montant_loyer') or '0'))
            jour = int(float(row.get('jour_echeance') or 1))
            if not (1 <= jour <= 31):
                raise ValueError("jour_echeance doit être entre 1 et 31")
            date_entree = _parse_date(row.get('date_entree')) or datetime.now().date()

            donnees = dict(
                bailleur=bailleur, nom=nom, prenom=prenom,
                telephone=(str(row.get('telephone') or '')).strip(),
                logement=(str(row.get('logement') or '')).strip(),
                montant_loyer=montant, jour_echeance=jour, date_entree=date_entree,
                solde_initial=Decimal(str(row.get('solde_initial') or '0')),
                date_debut_facturation=_parse_date(row.get('date_debut_facturation')),
            )
            apercu.append({'nom': nom, 'prenom': prenom, 'montant_loyer': str(montant)})
            if not dry_run:
                Locataire.objects.create(**donnees)
            crees += 1
        except (ValueError, InvalidOperation, KeyError) as e:
            erreurs.append({'ligne': i, 'message': str(e)})

    return {'crees': crees, 'erreurs': erreurs, 'apercu': apercu}
